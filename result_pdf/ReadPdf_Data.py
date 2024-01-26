import PyPDF2
import re
import openpyxl

wb_obj = openpyxl.load_workbook("data\\bi_tests2.xlsx")
sheet_obj = wb_obj.active
cell_obj = (sheet_obj.cell(row = 3, column = 5)).value


file = open(r'C:\fg-demo\result_pdf\result.pdf', 'rb')
reader = PyPDF2.PdfFileReader(file)
page1=reader.getPage(0)
temp_list = (page1.extractText().split('\n'))
page2=reader.getPage(1)
list2= (page2.extractText().split('\n'))
temp_list.extend(list2)
#print(temp_list)

list = [re.sub("[\x16]","ti",x) for x in temp_list]
print(list)
print("------------------------------------------------------------------------------------------------------------------------")
print("------------------------------------------------------------------------------------------------------------------------")

def compare_values(compare_option1, act_result1, exp_result1, step_no):
    print(f"Step No: {step_no}")
    print("Actual Result  : " +list[compare_option1]+f": {act_result1}")
    print("Expected Result: " +list[compare_option1]+f": {exp_result1}")
    if act_result1 == exp_result1:
        print("Expected and Actual values are matched")
    else:
        print("Expected and Actual values are not matched")
    print("-------------------------------------------------------------")

# step 1
compare_option = list.index('Name of the Life Assured')
act_result_index = (compare_option+1)
act_result = str(list[act_result_index].rstrip())
exp_result = str((sheet_obj.cell(row = 2, column = 8)).value)
compare_values(compare_option, act_result, exp_result, 1)
# step 2
compare_option = list.index('Age of Life Assured')
act_result_index = (compare_option+1)
act_result = int(list[act_result_index])
exp_result = int((sheet_obj.cell(row = 2, column = 7)).value)
compare_values(compare_option, act_result, exp_result, 2)
# step 3
compare_option = list.index('Pay')
act_result_index = (compare_option+1)
act_result_split = list[act_result_index]
act_result_firstWord = act_result_split.split()
act_result = act_result_firstWord[0]
exp_result = (sheet_obj.cell(row = 2, column = 13)).value
compare_values(compare_option, act_result, exp_result, 3)
# step 4
compare_option = list.index('Policy Term (years)')
act_result_index = (compare_option+1)
act_result = int(list[act_result_index].rstrip())
exp_result = int((sheet_obj.cell(row = 2, column = 14)).value)
compare_values(compare_option, act_result, exp_result, 4)
# step 5
compare_option = list.index('Annualized Premium')
act_result_index = (compare_option+1)
act_result = list[act_result_index].rstrip()
exp_result = (sheet_obj.cell(row = 2, column = 22)).value
compare_values(compare_option, act_result, exp_result, 5)
# step 6
compare_option = list.index('Accidental Death Sum Assured (Rs.)')
act_result_index = (compare_option+1)
act_result_R_Comma = list[act_result_index].rstrip()
act_result = int(act_result_R_Comma.replace(",", ""))
exp_result = int((sheet_obj.cell(row = 2, column = 20)).value)
compare_values(compare_option, act_result, exp_result, 6)
# step 7
compare_option = list.index('Sum Assured on Death ( Rs.)')
act_result_index = (compare_option+1)
act_result_R_Comma = list[act_result_index].rstrip()
act_result = int(act_result_R_Comma.replace(",", ""))
exp_result = int((sheet_obj.cell(row = 2, column = 12)).value)
compare_values(compare_option, act_result, exp_result, 7)
# step 8
compare_option_index = list.index('Policy Option')
act_result_index = list[compare_option_index + 1]
# a = 'Opon 2: Extra Life Cover (Life Cover with Accidental Death Beneﬁt)'
b =act_result_index.split("(")
c = b[0].split(":")
act_result_R_space = c[1].lstrip()
act_result = act_result_R_space.rstrip()
exp_result = (sheet_obj.cell(row = 2, column = 11)).value
compare_values(compare_option_index, act_result, exp_result, 8)