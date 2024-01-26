import pytest
from openpyxl import load_workbook
import json
import globals
from datetime import datetime
import os

class TestDriver:
    def main():
        globals.initialize()
        
        wb_suites = load_workbook("data\\test_suites.xlsx") # The main test suites file
        sheet_suites = wb_suites.active
        for test_suite in sheet_suites.iter_rows(min_row=2, values_only=True):
            if test_suite[2] == "y": # checking for the suite run condition
                wb_test_suite = load_workbook(test_suite[1]) # loading the test suite file
                sheet_test_suite = wb_test_suite.active
                for test_module in sheet_test_suite.iter_rows(min_row=2, values_only=True):
                    if test_module[2] == "y": # checking for the module run condition
                        wb_test_file = load_workbook(test_module[1]) # loading the test module file
                        sheet_test_file = wb_test_file.active
                        for test in sheet_test_file.iter_rows(min_row=2, values_only=True):
                            if test[2] == "y": # checking for the test run condition
                                # setting up the key as file name + test method name
                                globals.tests_to_run[f"tests\\{test[0]}::{test[1]}"] = [test[1], f"{test_module[1]}"] # assigning test data file to the test
                        wb_test_file.close()
                wb_test_suite.close()
        wb_suites.close()   
        dt_string = str(datetime.now().strftime("%d-%b-%Y(%H-%M-%S-%p)"))
        report_file_name = r"reports\report_" + dt_string + ".html"
        retcode = pytest.main(["--tb=no", "--capture=tee-sys", f"--html={report_file_name}", *globals.tests_to_run])
if __name__ == "__main__":
    TestDriver.main()
else:
    pass
