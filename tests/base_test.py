from selenium import webdriver
from openpyxl import load_workbook
import globals
from msedge.selenium_tools import Edge, EdgeOptions
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.common.keys import Keys

class BaseTest():
    def __init__(self):
        pass

    def setup_class(self):
        #FOR CHROME BROWSER
        #chrome_options = webdriver.ChromeOptions()
        #prefs = {"profile.default_content_setting_values.notifications" :2}
        #chrome_options.add_experimental_option("prefs", prefs)
        #self.driver = webdriver.Chrome(options=chrome_options)

        #executable path for chrome
        # self.driver = webdriver.Chrome()
        # self.driver = webdriver.Chrome(executable_path=r'drivers\\chromedriver.exe')

        #FOR EDGE BROWSER
        edge_options = webdriver.EdgeOptions()
        prefs = {"profile.default_content_setting_values.notifications": 2}
        edge_options.add_experimental_option("prefs", prefs)
        self.driver = webdriver.Edge(options=edge_options)

        #executable path for edge
        #edge_driver_path = 'drivers\\edgedriver.exe'
        #self.driver = webdriver.Edge(options=edge_options, executable_path=edge_driver_path) 
        self.driver.get("https://life.futuregenerali.in/")
        self.driver.maximize_window()

    def teardown_class(self):
        self.driver.close()
        self.driver.quit()

    def get_test_data(self, module_name, test_name):
        print(f'{module_name} : {test_name}')
        if f"tests\\{module_name}.py::{test_name}" in globals.tests_to_run:
            test_data_file = globals.tests_to_run[f"tests\\{module_name}.py::{test_name}"][1]
            wb_test_data = load_workbook(test_data_file)
            sheet_test_data = wb_test_data.active
            for test_data in sheet_test_data.iter_rows(min_row=2, values_only=True):
                if test_data[1] == test_name:
                    break
            wb_test_data.close()
            return test_data
            
    def chk_step_status(self, err_msg, step_num, step_desc, **kwargs):
        if len(err_msg) > 0:
            if "exclude_err_msg" in kwargs and kwargs["exclude_err_msg"] in err_msg:
                print(f"Step {str(step_num)}: Pass - {step_desc}")
                return True
            else:
                print(f"Step {str(step_num)}: Fail - {step_desc}")
                return False
        else:
            print(f"Step {str(step_num)}: Pass - {step_desc}")
            return True
    def test_step_status_log(self, step_status, step_num, step_desc):
        if step_status == 'p':
            print(f"Step {str(step_num)}: Pass - {step_desc}")
        elif step_status == 'f':
            print(f"Step {str(step_num)}: Fail - {step_desc}")
        elif step_status == 'e':
            print(f"Step {str(step_num)}: Exception - {step_desc}")

    
    def chk_alert_text(self):
        pass
